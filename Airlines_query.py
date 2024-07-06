# Please note that Barchart plotting are reffered from https://matplotlib.org
# Excel Reading is referred from https://www.geeksforgeeks.org/reading-excel-file-using-python/
# Excel sheet MCC codes are referred from https://www.dm.usda.gov/procurement/ccsc/docs/pcref/MCCandBOCCcrosswalk1-27-10.xls
# Basic code for MongoDB Database local connection
import matplotlib.pyplot as plt
import pymongo as pymongo
from pymongo import MongoClient
import numpy as np
client = MongoClient()
client = MongoClient("localhost", 27017)
mydatabase = client.project
collection = mydatabase.CS532

#Excel imports 
import openpyxl
path = "./MCC.xlsm"
Excel_sheet_loc = openpyxl.load_workbook(path)
Excel_obj = Excel_sheet_loc.active
row = Excel_obj.max_row
column = Excel_obj.max_column
Main_AIR_MCC = []
Main_AIR_names = []

# Query for getting distinct MCC codes available in the DB
Query_distinct_mcc = collection.distinct(
    "MCC"
)

# Saving MCC as integers for searching in excel and filtering out airlines 
MCC_int = [int(i) for i in Query_distinct_mcc]

# Reading Excel sheet and filtering out Airlines names 
for i in range(1, row + 1):
    Var_read_row = Excel_obj.cell(row=i, column=1)
    Var_read_col = Excel_obj.cell(row=i, column=2)
    if Var_read_row.value in MCC_int:
        if ("AIRLINE" in Var_read_col.value or "AIRWAY" in Var_read_col.value or " AIR " in Var_read_col.value):
            Main_AIR_MCC.append(Var_read_row.value)
            Main_AIR_names.append(Var_read_col.value)
            print("Printing Filtered Airlines", Var_read_row.value, Var_read_col.value)
            print("Please wait Plotting is in progress")
##
##
#Query-3(Part-1)
##
##
# Variables for plotting the graph
count_list_online = []
count_list_swipe = []
count_list_chip = []
year = []

match_mcc = {
    "$match": {
        "MCC": {
             "$in": [str(i) for i in Main_AIR_MCC]
        }
    }
    }
new_db = {
    "$out": {
        "db": "Airlines_db",
        "coll": "Airlines_coll"
    }  
}

pipeline = [
    match_mcc,
    new_db,
    ]

# Outing new database and executing our query on subDB
collection.aggregate(pipeline)
for k in range(2000, 2021, 1):
    match_year = {
        "$match": {
            "Year": str(k),
        }
    }
    match_online = {
        "$match": {
            "Use Chip": "Online Transaction"
        }
    }
    match_swipe = {
        "$match": {
            "Use Chip": "Swipe Transaction"
        }
    }
    match_chip = {
        "$match": {
            "Use Chip": "Chip Transaction"
        }
    }
    pipeline1 = [
        match_year,
        match_online,
    ]
    pipeline2 = [
        match_year,
        match_swipe,
    ]
    pipeline3 = [
        match_year,
        match_chip,
    ]

    # aggregating results and saving in python arrays
    year.append(k)
    results_online = mydatabase.Airlines_db.aggregate(pipeline1)
    results_swipe = mydatabase.Airlines_db.aggregate(pipeline2)
    results_chip = mydatabase.Airlines_db.aggregate(pipeline3)
    count1 = 0
    count2 = 0
    count3 = 0
    for x in results_online:
        count1 = count1 + 1
    for x in results_swipe:
        count2 = count2 + 1
    for x in results_chip:
        count3 = count2 + 1

    count_list_online.append(count1)
    count_list_swipe.append(count2)
    count_list_chip.append(count3)

labels = year
x1 = count_list_online
y = count_list_swipe
z = count_list_chip
barWidth = 0.25
bars1 = x1
bars2 = y
bars3 = z
r1 = np.arange(len(bars1))
r2 = [x + barWidth for x in r1]
r3 = [x + barWidth for x in r2]
plt.bar(r1, x1, color='#d41616', width=barWidth, edgecolor='white', label='Online Transaction')
plt.bar(r2, y, color='#405bad', width=barWidth, edgecolor='white', label='Swipe Transaction')
plt.bar(r3, z, color='#6ee039', width=barWidth, edgecolor='white', label='Chip Transaction')
plt.xlabel('group', fontweight='bold')
plt.xticks([r + barWidth for r in range(len(x1))], labels)
plt.legend()
plt.show()

##
##
#Query-3(Part-2)
##
##
# Variables declaration for plotting
count_list_chip = []
Airlines_dict = []
MCC_got = []
Cost_total = []
year = []
for k in range(2020,2021):
    match_year = {
        "$match": {
            "Year": str(k),
        }
    }
    for r in range(0,len(Main_AIR_MCC)):
        match_airline = {
            "$match": {
                "MCC": str(Main_AIR_MCC[r])
            }
        }
        project_cost = {
            "$project": {
            "MCC": 1,
            "cost_split": {
                "$split": [
                "$Amount", {
                    "$literal": "$"
                }
            ]
            }
        }
        }
        project_split_cost = {
            "$project": {
                "_id": 0,
                "MCC": 1,
                "cost_split": 1
        }   
        }
        unwind_only_cost = {
            "$unwind": {
                "path": "$cost_split",
                "includeArrayIndex": "arrayIndex"
        }   
        }
        include_array_0 = {
            "$match": {
            "arrayIndex": 1
            }
        }

        group_cost = {
            "$group" : {
                        "_id" : "$MCC",
                        "Total_cost" : {
                            "$sum" : {
                                "$convert": {
                                    "input": "$cost_split",
                                    "to": "double",
                                    "onError": "An error occurred",
                                }
                            }
                        }
                    }
        }
        limit = {
            "$limit": 10
        }
        pipeline1 = [
            match_year,
            match_airline,
            project_cost,
            project_split_cost,
            unwind_only_cost,
            include_array_0,
            group_cost,
            
        ]
        year.append(k)
        results = mydatabase.Airlines_db.aggregate(pipeline1)
        Airlines_dict = []
        for x in results:
            Airlines_dict.append(x)
       

        for dct in Airlines_dict: 
            MCC_got.append(dct['_id'])
            Cost_total.append(dct['Total_cost'])

# Plotting BarGraph
    x = [int(i) for i in MCC_got]
    y = [int(i) for i in Cost_total]
    zip = ["" for x in range(len(Main_AIR_names))]
    for i in range(0,len(Main_AIR_names)):
        zip[i] = str(x[i]) + '(' + Main_AIR_names[i] + ')'
    Yticks_set = zip
    set_airline = np.arange(len(Yticks_set))
    Main_vert_bar=y
    plt.barh(set_airline,Main_vert_bar,color='#4044ad',edgecolor='black')
    plt.yticks(set_airline, Yticks_set, fontsize=6)
    plt.xlabel('Total Profit', fontsize=16)
    plt.ylabel('Airlines', fontsize=6)
    plt.title('Profit of airlines during year 2020',fontsize=20)
    plt.show()
    


    

