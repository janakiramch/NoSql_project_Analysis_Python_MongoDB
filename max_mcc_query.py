# Please note that Barchart plotting are reffered from https://matplotlib.org
# Excel Reading is referred from https://www.geeksforgeeks.org/reading-excel-file-using-python/
# Excel sheet MCC codes are referred from https://www.dm.usda.gov/procurement/ccsc/docs/pcref/MCCandBOCCcrosswalk1-27-10.xls
# Basic code for MongoDB Database local connection
import pymongo as pymongo
from pymongo import MongoClient

# Excel imports for reading excel file
import openpyxl
import matplotlib.pyplot as plt
path = "./MCC.xlsm"
Excel_sheet_loc = openpyxl.load_workbook(path)
Excel_obj = Excel_sheet_loc.active
row = Excel_obj.max_row
column = Excel_obj.max_column

# Connecting to MongoDB by pymongo
client = MongoClient()
client = MongoClient("localhost", 27017)
mydatabase = client.project
collection = mydatabase.CS532

# Variables for plotting graphs
States_dict = []
States = []
CA_count = []
Mcc_save =[]
cnt =[]

# Query-1
#Getting Top States/countries with Most Transactions in 2-decades
results = collection.aggregate(
    [
        {
            "$match": {
                "Merchant State": {
                    "$ne": ""
                }
            }
        },

        {
            "$match" : {
                "Is Fraud?" : "No"
            }
        }, 

        {
            "$group" : {
                "_id" : "$Merchant State",
                "count" : {
                    "$sum" : 1.0
                }
            }
        },

        {
            "$sort" : {
                "count" : -1.0
            }
        },

        {
        "$limit" : 1
        },

        {
            "$group": {
                "_id": "$_id"
            }
        }, 
    ] 
)
# extracting States from above query result
for x in results:
    States_dict.append(x)
# Moving results aggreagte to States array
for dct in States_dict: 
    States.append(dct['_id'])
    print("State with highest transactions", States)
    print("Please wait Plotting in progress")
# looping all years for plotting the graph
for u in range(0, len(States)):
      for t in range (2000,2020, 2):
        extract_state = {
                    "$match": {
                        "Merchant State" : str(States[u])
                    }
                }
        extract_year = {
                    "$match": {
                        "Year": str(t)
                    }
                }
        most_oocuring_mcc = {
                    "$group" : {
                        "_id" : "$MCC",
                        "count" : {
                            "$sum" : 1.0
                        }
                    }
        }
        sort = {
            "$sort" : {
                        "count" : -1.0
                    },
        }
        limit = {
                    "$limit" : 1
                }

        pipeline = [
            extract_state,
            extract_year,
            most_oocuring_mcc,
            sort,
            limit,
        ]

        CA_Dict = []
        CA_mcc = []
        CA_result = collection.aggregate(pipeline)
        for x in CA_result:
            CA_Dict.append(x)

        for dct in CA_Dict: 
            CA_mcc.append(dct['_id'])
            CA_count.append(dct['count'])

# Reading excel sheet to get the values of MCC codes 
# Excel sheet is downloaded from official US MCC website      
        for k in range(0, len(CA_mcc)):
            for i in range(1, row + 1):
                Var_read_row = Excel_obj.cell(row=i, column=1)
                Var_read_col = Excel_obj.cell(row=i, column=2)
                if str(Var_read_row.value) in CA_mcc[k]:
                    Mcc_save.append(Var_read_col.value)

# Plotting Pie chart
legend_info = list(range(2000, 2020, 2))
converted_pie = ["" for x in range(len(legend_info))]
count_pie = [int(i) for i in CA_count]
count_pie_string = CA_count
bar_labels = Mcc_save  

# zipping MCC and Years for legend construction
for i in range(0, len(legend_info)):
    converted_pie[i] = str(legend_info[i]) + '(' + bar_labels[i] + ')'
pie_out = converted_pie
Data = count_pie
plt.pie(Data, labels = Data,counterclock=False, shadow=True)
plt.title("California state chart on spending sectors")
plt.legend(pie_out , bbox_to_anchor=(1.02,1), loc ="upper left", borderaxespad=0, prop = {'size': 7.5})
plt.show()